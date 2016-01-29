#!usr/bin/env

#sendDuesRemainders.py - Sends emails based on payment status in spreadsheet.
# open the spreadsheet and get the latest dues status

import openpyxl,smtplib, sys, getpass

#open the spreadsheet and get the latest dues status
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

lastCol = sheet.max_column
print(lastCol)

latestMonth = sheet.cell(row=1,column = lastCol).value
print(latestMonth)

#check each member's payment status
unpaidMembers = {}
for r in range(2,sheet.max_column):
	payment = sheet.cell(row = r,column=lastCol).value
	if payment != 'paid':
		name = sheet.cell(row = r,column = 1).value
		print(name)
		email = sheet.cell(row=r,column=2).value
		unpaidMembers[name] = email

#login to email account
smtpObj = smtplib.SMTP('smtp-mail.outlook.com',587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login('foysal.eee.bd@outlook.com',getpass.getpass('Password:'))

#send out remainder emails

for name,email in unpaidMembers.items():
	body = "Subject: %s dues unpaid.\n Dear %s,\nRecords show that you have noy paid dues for %s. Please make this payment as soon as possible. Thank you!"%(latestMonth,name,latestMonth)

	print('Sending email to %s...'%email)
	sendmailStatus = smtpObj.sendmail('foysal.eee.bd@outlook.com',email,body)

	if sendmailStatus != {}:
		print('There was a problem sending email to %s: %s'%(email,sendmailStatus))

smtpObj.quit()		





